#!/usr/bin/env python3
"""
MyRoofGenius Brand Template System
Ensures consistent branding across all digital products
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@dataclass
class BrandColors:
    """MyRoofGenius brand colors"""
    primary_blue: str = "#1E3A8A"      # Deep blue
    secondary_blue: str = "#3B82F6"    # Bright blue
    accent_green: str = "#10B981"      # Success green
    text_dark: str = "#1F2937"         # Dark gray
    text_light: str = "#6B7280"        # Medium gray
    background: str = "#FFFFFF"        # White
    background_alt: str = "#F9FAFB"    # Light gray
    error_red: str = "#EF4444"         # Error red
    warning_yellow: str = "#F59E0B"    # Warning yellow

@dataclass
class BrandFonts:
    """MyRoofGenius brand fonts"""
    heading_family: str = "Inter"
    body_family: str = "Inter"
    mono_family: str = "JetBrains Mono"
    heading_size: int = 24
    subheading_size: int = 18
    body_size: int = 12
    small_size: int = 10

class BrandTemplateSystem:
    """Manages brand templates and styling for all products"""
    
    def __init__(self):
        self.colors = BrandColors()
        self.fonts = BrandFonts()
        self.logo_path = "/home/mwwoodworth/code/assets/myroofgenius_logo.png"
        self.templates_dir = Path("/home/mwwoodworth/code/brand_templates")
        self.templates_dir.mkdir(exist_ok=True)
        
    def create_excel_template(self, template_name: str, sheet_configs: List[Dict[str, Any]]) -> str:
        """Create a branded Excel template"""
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets based on config
        for config in sheet_configs:
            sheet_name = config.get('name', 'Sheet')
            sheet_type = config.get('type', 'data')
            
            ws = wb.create_sheet(title=sheet_name)
            
            if sheet_type == 'cover':
                self._create_excel_cover_sheet(ws)
            elif sheet_type == 'instructions':
                self._create_excel_instructions_sheet(ws, config.get('content', {}))
            elif sheet_type == 'data':
                self._create_excel_data_sheet(ws, config.get('columns', []), config.get('sample_data', []))
            elif sheet_type == 'calculator':
                self._create_excel_calculator_sheet(ws, config.get('calculations', {}))
            elif sheet_type == 'dashboard':
                self._create_excel_dashboard_sheet(ws, config.get('metrics', []))
        
        # Save template
        output_path = self.templates_dir / f"{template_name}.xlsx"
        wb.save(output_path)
        
        logger.info(f"Created Excel template: {output_path}")
        return str(output_path)
    
    def _create_excel_cover_sheet(self, ws):
        """Create branded cover sheet for Excel"""
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        
        # Add logo placeholder
        ws['B2'] = "MyRoofGenius"
        ws['B2'].font = Font(name=self.fonts.heading_family, size=28, bold=True, 
                            color=self.colors.primary_blue.replace('#', ''))
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add title
        ws.merge_cells('B5:C5')
        ws['B5'] = "Professional Roofing Business Template"
        ws['B5'].font = Font(name=self.fonts.heading_family, size=20, bold=True)
        ws['B5'].alignment = Alignment(horizontal='center')
        
        # Add subtitle
        ws.merge_cells('B7:C7')
        ws['B7'] = "Streamline Your Business Operations"
        ws['B7'].font = Font(name=self.fonts.body_family, size=14, 
                            color=self.colors.text_light.replace('#', ''))
        ws['B7'].alignment = Alignment(horizontal='center')
        
        # Add feature list
        features = [
            "✓ Easy-to-use formulas and calculations",
            "✓ Professional formatting and design",
            "✓ Customizable for your business needs",
            "✓ Includes examples and instructions",
            "✓ Compatible with Excel 2016 and newer"
        ]
        
        start_row = 10
        for i, feature in enumerate(features):
            ws[f'B{start_row + i}'] = feature
            ws[f'B{start_row + i}'].font = Font(name=self.fonts.body_family, size=12)
        
        # Add copyright
        ws.merge_cells('B20:C20')
        ws['B20'] = f"© {datetime.now().year} MyRoofGenius. All rights reserved."
        ws['B20'].font = Font(name=self.fonts.body_family, size=10, 
                             color=self.colors.text_light.replace('#', ''))
        ws['B20'].alignment = Alignment(horizontal='center')
        
        # Add borders and styling
        self._apply_excel_styling(ws)
    
    def _create_excel_instructions_sheet(self, ws, content: Dict[str, Any]):
        """Create instructions sheet with professional formatting"""
        # Header
        ws['A1'] = "Instructions"
        ws['A1'].font = Font(name=self.fonts.heading_family, size=24, bold=True,
                            color=self.colors.primary_blue.replace('#', ''))
        
        # Section styling
        current_row = 3
        
        sections = content.get('sections', [
            {
                "title": "Getting Started",
                "content": [
                    "1. Save a copy of this template for your business",
                    "2. Update the company information on the Settings sheet",
                    "3. Begin entering your data in the appropriate sheets",
                    "4. Formulas will calculate automatically"
                ]
            },
            {
                "title": "Features",
                "content": [
                    "• Automatic calculations for all key metrics",
                    "• Professional invoice and estimate generation",
                    "• Customer tracking and management",
                    "• Financial reporting and analytics"
                ]
            }
        ])
        
        for section in sections:
            # Section title
            ws[f'A{current_row}'] = section['title']
            ws[f'A{current_row}'].font = Font(name=self.fonts.heading_family, size=16, bold=True)
            current_row += 2
            
            # Section content
            for item in section['content']:
                ws[f'B{current_row}'] = item
                ws[f'B{current_row}'].font = Font(name=self.fonts.body_family, size=12)
                current_row += 1
            
            current_row += 2
        
        # Support information
        ws[f'A{current_row}'] = "Need Help?"
        ws[f'A{current_row}'].font = Font(name=self.fonts.heading_family, size=14, bold=True)
        current_row += 1
        
        ws[f'B{current_row}'] = "Visit support.myroofgenius.com or email support@myroofgenius.com"
        ws[f'B{current_row}'].font = Font(name=self.fonts.body_family, size=12,
                                         color=self.colors.secondary_blue.replace('#', ''))
    
    def _create_excel_data_sheet(self, ws, columns: List[Dict[str, str]], sample_data: List[List[Any]]):
        """Create a data entry sheet with proper formatting"""
        # Header row
        header_fill = PatternFill(start_color=self.colors.primary_blue.replace('#', ''),
                                 end_color=self.colors.primary_blue.replace('#', ''),
                                 fill_type='solid')
        
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column.get('name', f'Column {col_idx}')
            cell.font = Font(name=self.fonts.body_family, size=12, bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column width
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column.get('width', 15)
        
        # Add sample data with alternating row colors
        light_fill = PatternFill(start_color=self.colors.background_alt.replace('#', ''),
                                end_color=self.colors.background_alt.replace('#', ''),
                                fill_type='solid')
        
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = Font(name=self.fonts.body_family, size=11)
                
                if row_idx % 2 == 0:
                    cell.fill = light_fill
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
    
    def _create_excel_calculator_sheet(self, ws, calculations: Dict[str, Any]):
        """Create a calculator sheet with formulas"""
        ws['A1'] = calculations.get('title', 'Calculator')
        ws['A1'].font = Font(name=self.fonts.heading_family, size=20, bold=True,
                            color=self.colors.primary_blue.replace('#', ''))
        
        # Input section
        ws['A3'] = "Input Values"
        ws['A3'].font = Font(name=self.fonts.heading_family, size=14, bold=True)
        
        input_start_row = 5
        inputs = calculations.get('inputs', [])
        
        for i, input_field in enumerate(inputs):
            row = input_start_row + i
            ws[f'A{row}'] = input_field['label']
            ws[f'A{row}'].font = Font(name=self.fonts.body_family, size=12)
            
            ws[f'C{row}'] = input_field.get('default', 0)
            ws[f'C{row}'].font = Font(name=self.fonts.body_family, size=12, bold=True)
            ws[f'C{row}'].fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
            
            if 'unit' in input_field:
                ws[f'D{row}'] = input_field['unit']
                ws[f'D{row}'].font = Font(name=self.fonts.body_family, size=11,
                                         color=self.colors.text_light.replace('#', ''))
        
        # Results section
        results_start_row = input_start_row + len(inputs) + 3
        ws[f'A{results_start_row}'] = "Calculated Results"
        ws[f'A{results_start_row}'].font = Font(name=self.fonts.heading_family, size=14, bold=True)
        
        results = calculations.get('results', [])
        for i, result in enumerate(results):
            row = results_start_row + 2 + i
            ws[f'A{row}'] = result['label']
            ws[f'A{row}'].font = Font(name=self.fonts.body_family, size=12)
            
            # Add formula
            ws[f'C{row}'] = result.get('formula', '=0')
            ws[f'C{row}'].font = Font(name=self.fonts.body_family, size=12, bold=True)
            ws[f'C{row}'].fill = PatternFill(start_color=self.colors.accent_green.replace('#', ''),
                                            end_color=self.colors.accent_green.replace('#', ''),
                                            fill_type='solid')
            
            if 'unit' in result:
                ws[f'D{row}'] = result['unit']
    
    def _create_excel_dashboard_sheet(self, ws, metrics: List[Dict[str, Any]]):
        """Create a dashboard sheet with visual metrics"""
        ws['A1'] = "Business Dashboard"
        ws['A1'].font = Font(name=self.fonts.heading_family, size=24, bold=True,
                            color=self.colors.primary_blue.replace('#', ''))
        
        # KPI Cards
        card_positions = [
            ('B3', 'C5'),  # Card 1
            ('E3', 'F5'),  # Card 2
            ('H3', 'I5'),  # Card 3
            ('B7', 'C9'),  # Card 4
            ('E7', 'F9'),  # Card 5
            ('H7', 'I9'),  # Card 6
        ]
        
        for i, (metric, pos) in enumerate(zip(metrics[:6], card_positions)):
            start_cell, end_cell = pos
            ws.merge_cells(f'{start_cell}:{end_cell}')
            
            # Add metric value
            ws[start_cell] = metric.get('value', '0')
            ws[start_cell].font = Font(name=self.fonts.heading_family, size=28, bold=True,
                                      color=self.colors.primary_blue.replace('#', ''))
            ws[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            
            # Add metric label below
            label_row = int(end_cell[1:]) + 1
            label_cell = f'{start_cell[0]}{label_row}'
            ws.merge_cells(f'{label_cell}:{end_cell[0]}{label_row}')
            ws[label_cell] = metric.get('label', 'Metric')
            ws[label_cell].font = Font(name=self.fonts.body_family, size=12,
                                      color=self.colors.text_light.replace('#', ''))
            ws[label_cell].alignment = Alignment(horizontal='center')
    
    def _apply_excel_styling(self, ws):
        """Apply consistent styling to Excel worksheet"""
        # Set default font for entire sheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and not cell.font.bold:
                    cell.font = Font(name=self.fonts.body_family, size=self.fonts.body_size)
    
    def create_pdf_template(self, template_name: str, content: Dict[str, Any]) -> str:
        """Create a branded PDF template"""
        output_path = self.templates_dir / f"{template_name}.pdf"
        
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        
        # Create custom styles
        styles = self._create_pdf_styles()
        
        # Build document content
        story = []
        
        # Add cover page
        story.extend(self._create_pdf_cover(content.get('title', 'Document'), 
                                          content.get('subtitle', ''), styles))
        
        # Add table of contents if needed
        if content.get('include_toc', True):
            story.extend(self._create_pdf_toc(content.get('sections', []), styles))
        
        # Add content sections
        for section in content.get('sections', []):
            story.extend(self._create_pdf_section(section, styles))
        
        # Build PDF
        doc.build(story)
        
        logger.info(f"Created PDF template: {output_path}")
        return str(output_path)
    
    def _create_pdf_styles(self) -> Dict[str, ParagraphStyle]:
        """Create custom PDF styles matching brand"""
        styles = getSampleStyleSheet()
        
        # Custom title style
        styles.add(ParagraphStyle(
            name='BrandTitle',
            parent=styles['Title'],
            fontSize=28,
            textColor=HexColor(self.colors.primary_blue),
            spaceAfter=30,
            alignment=1  # Center
        ))
        
        # Custom heading style
        styles.add(ParagraphStyle(
            name='BrandHeading',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=HexColor(self.colors.primary_blue),
            spaceAfter=12,
            spaceBefore=24
        ))
        
        # Custom subheading style
        styles.add(ParagraphStyle(
            name='BrandSubheading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=HexColor(self.colors.text_dark),
            spaceAfter=10,
            spaceBefore=16
        ))
        
        # Custom body style
        styles.add(ParagraphStyle(
            name='BrandBody',
            parent=styles['BodyText'],
            fontSize=12,
            textColor=HexColor(self.colors.text_dark),
            spaceAfter=12,
            leading=18
        ))
        
        return styles
    
    def _create_pdf_cover(self, title: str, subtitle: str, styles: Dict) -> List:
        """Create PDF cover page"""
        story = []
        
        # Add space at top
        story.append(Spacer(1, 2*inch))
        
        # Add title
        story.append(Paragraph(title, styles['BrandTitle']))
        
        # Add subtitle
        if subtitle:
            story.append(Paragraph(subtitle, styles['BrandSubheading']))
        
        # Add space
        story.append(Spacer(1, 1*inch))
        
        # Add branding text
        story.append(Paragraph("MyRoofGenius", styles['BrandHeading']))
        story.append(Paragraph("Professional Business Solutions for Roofing Contractors", 
                             styles['BrandBody']))
        
        # Page break
        story.append(PageBreak())
        
        return story
    
    def _create_pdf_toc(self, sections: List[Dict[str, Any]], styles: Dict) -> List:
        """Create table of contents"""
        story = []
        
        story.append(Paragraph("Table of Contents", styles['BrandHeading']))
        story.append(Spacer(1, 0.5*inch))
        
        for i, section in enumerate(sections, 1):
            toc_entry = f"{i}. {section.get('title', 'Section')}"
            story.append(Paragraph(toc_entry, styles['BrandBody']))
        
        story.append(PageBreak())
        
        return story
    
    def _create_pdf_section(self, section: Dict[str, Any], styles: Dict) -> List:
        """Create PDF section"""
        story = []
        
        # Section title
        story.append(Paragraph(section.get('title', 'Section'), styles['BrandHeading']))
        
        # Section content
        for paragraph in section.get('content', []):
            if isinstance(paragraph, dict):
                if paragraph.get('type') == 'subheading':
                    story.append(Paragraph(paragraph['text'], styles['BrandSubheading']))
                elif paragraph.get('type') == 'list':
                    for item in paragraph.get('items', []):
                        story.append(Paragraph(f"• {item}", styles['BrandBody']))
                else:
                    story.append(Paragraph(paragraph.get('text', ''), styles['BrandBody']))
            else:
                story.append(Paragraph(str(paragraph), styles['BrandBody']))
        
        # Add spacing after section
        story.append(Spacer(1, 0.5*inch))
        
        return story
    
    def create_notion_template_structure(self, template_name: str, 
                                       structure: Dict[str, Any]) -> Dict[str, Any]:
        """Create Notion template structure with branding"""
        notion_template = {
            "name": template_name,
            "branding": {
                "icon": "🏠",  # Roof/house icon
                "cover": {
                    "type": "gradient",
                    "gradient": {
                        "start": self.colors.primary_blue,
                        "end": self.colors.secondary_blue
                    }
                }
            },
            "properties": {},
            "content": [],
            "views": []
        }
        
        # Add branded header
        notion_template['content'].append({
            "type": "heading_1",
            "text": structure.get('title', template_name),
            "color": "blue"
        })
        
        # Add description
        if 'description' in structure:
            notion_template['content'].append({
                "type": "text",
                "text": structure['description']
            })
        
        # Add database properties
        if 'database_properties' in structure:
            for prop in structure['database_properties']:
                notion_template['properties'][prop['name']] = {
                    "type": prop['type'],
                    "options": prop.get('options', [])
                }
        
        # Add views
        if 'views' in structure:
            for view in structure['views']:
                notion_template['views'].append({
                    "type": view['type'],
                    "name": view['name'],
                    "filters": view.get('filters', []),
                    "sorts": view.get('sorts', [])
                })
        
        # Save template structure
        output_path = self.templates_dir / f"{template_name}_notion.json"
        with open(output_path, 'w') as f:
            json.dump(notion_template, f, indent=2)
        
        logger.info(f"Created Notion template structure: {output_path}")
        return notion_template
    
    def apply_branding_to_existing_file(self, file_path: str, file_type: str) -> str:
        """Apply branding to an existing file"""
        file_path = Path(file_path)
        
        if file_type == 'excel' and file_path.suffix in ['.xlsx', '.xls']:
            return self._brand_excel_file(file_path)
        elif file_type == 'pdf' and file_path.suffix == '.pdf':
            logger.warning("PDF rebranding requires reconstruction")
            return str(file_path)
        else:
            logger.warning(f"Unsupported file type for branding: {file_type}")
            return str(file_path)
    
    def _brand_excel_file(self, file_path: Path) -> str:
        """Apply branding to existing Excel file"""
        wb = openpyxl.load_workbook(file_path)
        
        # Apply branding to each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Apply header styling
            if ws.max_row > 0:
                header_fill = PatternFill(
                    start_color=self.colors.primary_blue.replace('#', ''),
                    end_color=self.colors.primary_blue.replace('#', ''),
                    fill_type='solid'
                )
                
                for cell in ws[1]:
                    if cell.value:
                        cell.font = Font(name=self.fonts.body_family, size=12, 
                                       bold=True, color='FFFFFF')
                        cell.fill = header_fill
            
            # Apply general styling
            self._apply_excel_styling(ws)
        
        # Save branded version
        branded_path = file_path.parent / f"{file_path.stem}_branded{file_path.suffix}"
        wb.save(branded_path)
        
        logger.info(f"Applied branding to Excel file: {branded_path}")
        return str(branded_path)
    
    def generate_brand_guidelines_document(self) -> str:
        """Generate a brand guidelines document"""
        content = {
            "title": "MyRoofGenius Brand Guidelines",
            "subtitle": "Ensuring Consistency Across All Digital Products",
            "sections": [
                {
                    "title": "Brand Colors",
                    "content": [
                        {"type": "subheading", "text": "Primary Colors"},
                        f"Primary Blue: {self.colors.primary_blue}",
                        f"Secondary Blue: {self.colors.secondary_blue}",
                        f"Accent Green: {self.colors.accent_green}",
                        {"type": "subheading", "text": "Text Colors"},
                        f"Dark Text: {self.colors.text_dark}",
                        f"Light Text: {self.colors.text_light}",
                        {"type": "subheading", "text": "Background Colors"},
                        f"Primary Background: {self.colors.background}",
                        f"Alternative Background: {self.colors.background_alt}"
                    ]
                },
                {
                    "title": "Typography",
                    "content": [
                        {"type": "subheading", "text": "Font Families"},
                        f"Headings: {self.fonts.heading_family}",
                        f"Body Text: {self.fonts.body_family}",
                        f"Monospace: {self.fonts.mono_family}",
                        {"type": "subheading", "text": "Font Sizes"},
                        f"Main Heading: {self.fonts.heading_size}pt",
                        f"Subheading: {self.fonts.subheading_size}pt",
                        f"Body: {self.fonts.body_size}pt",
                        f"Small Text: {self.fonts.small_size}pt"
                    ]
                },
                {
                    "title": "Logo Usage",
                    "content": [
                        "The MyRoofGenius logo should appear on all digital products",
                        "Minimum size: 120px width for digital, 1 inch for print",
                        "Clear space: Maintain minimum 20px clearance around logo",
                        "Placement: Top-left corner or centered in header"
                    ]
                },
                {
                    "title": "Document Standards",
                    "content": [
                        {"type": "subheading", "text": "General Guidelines"},
                        "Use consistent margins (minimum 1 inch)",
                        "Maintain 1.5 line spacing for body text",
                        "Include page numbers on multi-page documents",
                        "Add copyright notice in footer",
                        {"type": "subheading", "text": "Professional Tone"},
                        "Use clear, concise language",
                        "Focus on value for roofing contractors",
                        "Include practical examples and use cases",
                        "Provide actionable insights and tools"
                    ]
                }
            ]
        }
        
        return self.create_pdf_template("brand_guidelines", content)

# Utility functions for creating specific product types
def create_roofing_estimate_template():
    """Create a professional roofing estimate template"""
    template_system = BrandTemplateSystem()
    
    sheet_configs = [
        {
            "name": "Cover",
            "type": "cover"
        },
        {
            "name": "Instructions",
            "type": "instructions",
            "content": {
                "sections": [
                    {
                        "title": "How to Use This Estimate Template",
                        "content": [
                            "1. Update your company information in the Settings sheet",
                            "2. Enter customer details in the Customer Info section",
                            "3. Add line items for materials and labor",
                            "4. The template will automatically calculate totals",
                            "5. Print or export to PDF for customer presentation"
                        ]
                    }
                ]
            }
        },
        {
            "name": "Estimate",
            "type": "data",
            "columns": [
                {"name": "Item Description", "width": 30},
                {"name": "Quantity", "width": 10},
                {"name": "Unit", "width": 10},
                {"name": "Unit Price", "width": 15},
                {"name": "Total", "width": 15}
            ],
            "sample_data": [
                ["Architectural Shingles", 25, "Squares", "$125.00", "=B2*D2"],
                ["Underlayment", 25, "Squares", "$45.00", "=B3*D3"],
                ["Ridge Cap", 100, "Lin. Ft.", "$3.50", "=B4*D4"],
                ["Labor - Installation", 25, "Squares", "$150.00", "=B5*D5"]
            ]
        },
        {
            "name": "Calculator",
            "type": "calculator",
            "calculations": {
                "title": "Roofing Calculator",
                "inputs": [
                    {"label": "Roof Area (sq ft)", "default": 2500, "unit": "sq ft"},
                    {"label": "Roof Pitch", "default": 6, "unit": "/12"},
                    {"label": "Waste Factor", "default": 10, "unit": "%"}
                ],
                "results": [
                    {"label": "Total Squares Needed", "formula": "=ROUNDUP((C5*(1+C7/100))/100,0)", "unit": "squares"},
                    {"label": "Estimated Material Cost", "formula": "=C11*125", "unit": "$"},
                    {"label": "Estimated Labor Cost", "formula": "=C11*150", "unit": "$"},
                    {"label": "Total Estimate", "formula": "=C12+C13", "unit": "$"}
                ]
            }
        }
    ]
    
    return template_system.create_excel_template("roofing_estimate_template", sheet_configs)

if __name__ == "__main__":
    # Initialize brand template system
    brand_system = BrandTemplateSystem()
    
    # Generate brand guidelines
    guidelines_path = brand_system.generate_brand_guidelines_document()
    logger.info(f"Brand guidelines created: {guidelines_path}")
    
    # Create sample roofing estimate template
    estimate_path = create_roofing_estimate_template()
    logger.info(f"Roofing estimate template created: {estimate_path}")
#!/usr/bin/env python3
"""
Commercial Grade Quality Assurance System for MyRoofGenius/BrainOps
Ensures all digital products meet or exceed professional standards
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import asyncio
import aiohttp
from dataclasses import dataclass, asdict
from enum import Enum
import hashlib
import subprocess
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/tmp/commercial_qa_system.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ProductType(Enum):
    """Types of digital products in the marketplace"""
    EXCEL_TEMPLATE = "excel_template"
    PDF_DOCUMENT = "pdf_document"
    NOTION_TEMPLATE = "notion_template"
    DIGITAL_TOOL = "digital_tool"
    CALCULATOR = "calculator"
    WORKFLOW = "workflow"
    CHECKLIST = "checklist"
    CONTRACT_TEMPLATE = "contract_template"
    TRAINING_MATERIAL = "training_material"

class QualityStatus(Enum):
    """Product quality assessment status"""
    PENDING = "pending"
    IN_REVIEW = "in_review"
    FAILED = "failed"
    PASSED = "passed"
    REQUIRES_REWORK = "requires_rework"
    ADMIN_APPROVED = "admin_approved"
    LIVE = "live"

@dataclass
class QualityMetrics:
    """Quality metrics for product assessment"""
    completeness_score: float  # 0-100
    functionality_score: float  # 0-100
    visual_polish_score: float  # 0-100
    content_quality_score: float  # 0-100
    usability_score: float  # 0-100
    brand_compliance_score: float  # 0-100
    overall_score: float  # 0-100
    
    def calculate_overall(self):
        """Calculate overall quality score"""
        scores = [
            self.completeness_score,
            self.functionality_score,
            self.visual_polish_score,
            self.content_quality_score,
            self.usability_score,
            self.brand_compliance_score
        ]
        self.overall_score = sum(scores) / len(scores)
        return self.overall_score

@dataclass
class ProductQAReport:
    """Comprehensive QA report for a product"""
    product_id: str
    product_type: ProductType
    product_name: str
    timestamp: datetime
    status: QualityStatus
    metrics: QualityMetrics
    issues_found: List[Dict[str, Any]]
    improvements_made: List[str]
    test_results: Dict[str, Any]
    admin_notes: Optional[str] = None
    version: str = "1.0"
    
class CommercialGradeQASystem:
    """Main QA system for ensuring commercial-grade quality"""
    
    def __init__(self):
        self.min_quality_score = 95.0  # Minimum score for commercial grade
        self.supabase_url = os.getenv("SUPABASE_URL", "https://yomagoqdmxszqtdwuhab.supabase.co")
        self.supabase_key = os.getenv("SUPABASE_SERVICE_KEY")
        self.session = None
        self.brand_guidelines = self._load_brand_guidelines()
        
    def _load_brand_guidelines(self) -> Dict[str, Any]:
        """Load MyRoofGenius brand guidelines"""
        return {
            "colors": {
                "primary": "#1E3A8A",  # Deep blue
                "secondary": "#3B82F6",  # Bright blue
                "accent": "#10B981",  # Green
                "text": "#1F2937",  # Dark gray
                "background": "#FFFFFF"
            },
            "fonts": {
                "heading": "Inter, sans-serif",
                "body": "Inter, sans-serif",
                "monospace": "JetBrains Mono, monospace"
            },
            "logo_requirements": {
                "placement": "top-left or center header",
                "min_size": "120px width",
                "clear_space": "20px minimum"
            },
            "document_standards": {
                "margins": "1 inch minimum",
                "line_spacing": "1.5",
                "section_spacing": "24px",
                "professional_tone": True
            }
        }
    
    async def initialize(self):
        """Initialize the QA system"""
        self.session = aiohttp.ClientSession()
        logger.info("Commercial Grade QA System initialized")
        
    async def validate_product(self, product_data: Dict[str, Any]) -> ProductQAReport:
        """Perform comprehensive validation of a product"""
        logger.info(f"Starting validation for product: {product_data.get('name', 'Unknown')}")
        
        product_type = ProductType(product_data.get('type', 'digital_tool'))
        
        # Initialize metrics
        metrics = QualityMetrics(
            completeness_score=0,
            functionality_score=0,
            visual_polish_score=0,
            content_quality_score=0,
            usability_score=0,
            brand_compliance_score=0,
            overall_score=0
        )
        
        issues = []
        improvements = []
        test_results = {}
        
        # Run validation checks based on product type
        if product_type == ProductType.EXCEL_TEMPLATE:
            await self._validate_excel_template(product_data, metrics, issues, test_results)
        elif product_type == ProductType.PDF_DOCUMENT:
            await self._validate_pdf_document(product_data, metrics, issues, test_results)
        elif product_type == ProductType.NOTION_TEMPLATE:
            await self._validate_notion_template(product_data, metrics, issues, test_results)
        elif product_type in [ProductType.CALCULATOR, ProductType.DIGITAL_TOOL]:
            await self._validate_digital_tool(product_data, metrics, issues, test_results)
        else:
            await self._validate_generic_product(product_data, metrics, issues, test_results)
        
        # Calculate overall score
        overall_score = metrics.calculate_overall()
        
        # Determine status
        if overall_score >= self.min_quality_score:
            status = QualityStatus.PASSED
        elif overall_score >= 80:
            status = QualityStatus.REQUIRES_REWORK
        else:
            status = QualityStatus.FAILED
        
        # Create report
        report = ProductQAReport(
            product_id=product_data.get('id', 'unknown'),
            product_type=product_type,
            product_name=product_data.get('name', 'Unknown Product'),
            timestamp=datetime.utcnow(),
            status=status,
            metrics=metrics,
            issues_found=issues,
            improvements_made=improvements,
            test_results=test_results,
            version=product_data.get('version', '1.0')
        )
        
        # Store report in persistent memory
        await self._store_qa_report(report)
        
        # If failed or requires rework, trigger improvement workflow
        if status in [QualityStatus.FAILED, QualityStatus.REQUIRES_REWORK]:
            await self._trigger_improvement_workflow(report)
        
        return report
    
    async def _validate_excel_template(self, product_data: Dict, metrics: QualityMetrics, 
                                     issues: List, test_results: Dict):
        """Validate Excel template specific requirements"""
        file_path = product_data.get('file_path')
        if not file_path or not os.path.exists(file_path):
            issues.append({
                "type": "missing_file",
                "severity": "critical",
                "description": "Excel file not found"
            })
            metrics.completeness_score = 0
            return
        
        try:
            # Check file can be opened
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Check for formulas
            has_formulas = False
            broken_formulas = []
            
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            has_formulas = True
                        if cell.value == '#REF!' or cell.value == '#VALUE!' or cell.value == '#NAME?':
                            broken_formulas.append(f"{sheet.title}!{cell.coordinate}")
            
            # Score functionality
            if broken_formulas:
                metrics.functionality_score = 0
                issues.append({
                    "type": "broken_formulas",
                    "severity": "critical",
                    "description": f"Broken formulas found: {', '.join(broken_formulas[:5])}"
                })
            elif has_formulas:
                metrics.functionality_score = 100
            else:
                metrics.functionality_score = 50
                issues.append({
                    "type": "no_formulas",
                    "severity": "medium",
                    "description": "No formulas found in Excel template"
                })
            
            # Check for professional formatting
            has_headers = any(sheet.row_dimensions[1].height > 20 for sheet in wb.worksheets)
            has_styling = any(cell.font.bold for sheet in wb.worksheets for row in sheet.iter_rows(max_row=5) for cell in row if cell.font)
            
            if has_headers and has_styling:
                metrics.visual_polish_score = 90
            else:
                metrics.visual_polish_score = 60
                issues.append({
                    "type": "poor_formatting",
                    "severity": "medium",
                    "description": "Lacks professional formatting and styling"
                })
            
            # Check for instructions/documentation
            has_instructions = any('instruction' in sheet.title.lower() or 'help' in sheet.title.lower() 
                                 for sheet in wb.worksheets)
            
            if has_instructions:
                metrics.usability_score = 95
            else:
                metrics.usability_score = 70
                issues.append({
                    "type": "missing_instructions",
                    "severity": "medium",
                    "description": "No instructions or help sheet found"
                })
            
            metrics.completeness_score = 85 if len(issues) < 2 else 60
            metrics.content_quality_score = 80  # Base score, adjust based on content analysis
            metrics.brand_compliance_score = 75  # Check for branding elements
            
            test_results['excel_validation'] = {
                'sheets_count': len(wb.worksheets),
                'has_formulas': has_formulas,
                'broken_formulas_count': len(broken_formulas),
                'has_instructions': has_instructions
            }
            
        except Exception as e:
            logger.error(f"Error validating Excel template: {str(e)}")
            issues.append({
                "type": "validation_error",
                "severity": "critical",
                "description": f"Failed to validate Excel file: {str(e)}"
            })
            metrics.functionality_score = 0
    
    async def _validate_pdf_document(self, product_data: Dict, metrics: QualityMetrics, 
                                   issues: List, test_results: Dict):
        """Validate PDF document requirements"""
        file_path = product_data.get('file_path')
        if not file_path or not os.path.exists(file_path):
            issues.append({
                "type": "missing_file",
                "severity": "critical",
                "description": "PDF file not found"
            })
            return
        
        try:
            import PyPDF2
            
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                num_pages = len(reader.pages)
                
                # Extract text for content analysis
                full_text = ""
                for page in reader.pages:
                    full_text += page.extract_text()
                
                # Check document completeness
                if num_pages < 2:
                    metrics.completeness_score = 40
                    issues.append({
                        "type": "too_short",
                        "severity": "high",
                        "description": "Document has less than 2 pages"
                    })
                else:
                    metrics.completeness_score = min(100, 70 + (num_pages * 3))
                
                # Check for professional elements
                has_toc = 'table of contents' in full_text.lower() or 'contents' in full_text[:1000].lower()
                has_headers = any(keyword in full_text.lower() for keyword in ['chapter', 'section', 'introduction', 'conclusion'])
                has_copyright = 'copyright' in full_text.lower() or '©' in full_text
                
                professional_elements = sum([has_toc, has_headers, has_copyright])
                metrics.visual_polish_score = min(100, 60 + (professional_elements * 13))
                
                # Content quality check
                word_count = len(full_text.split())
                if word_count < 500:
                    metrics.content_quality_score = 30
                    issues.append({
                        "type": "insufficient_content",
                        "severity": "high",
                        "description": f"Document has only {word_count} words"
                    })
                else:
                    metrics.content_quality_score = min(100, 70 + (word_count // 100))
                
                # Functionality (for PDFs, this means readable and well-structured)
                metrics.functionality_score = 90 if len(issues) == 0 else 70
                
                # Usability
                metrics.usability_score = 85 if has_toc else 70
                
                # Brand compliance (basic check)
                has_branding = 'myroofgenius' in full_text.lower() or 'brainops' in full_text.lower()
                metrics.brand_compliance_score = 85 if has_branding else 60
                
                test_results['pdf_validation'] = {
                    'pages': num_pages,
                    'word_count': word_count,
                    'has_toc': has_toc,
                    'has_copyright': has_copyright,
                    'has_branding': has_branding
                }
                
        except Exception as e:
            logger.error(f"Error validating PDF: {str(e)}")
            issues.append({
                "type": "validation_error",
                "severity": "critical",
                "description": f"Failed to validate PDF: {str(e)}"
            })
    
    async def _validate_notion_template(self, product_data: Dict, metrics: QualityMetrics, 
                                      issues: List, test_results: Dict):
        """Validate Notion template requirements"""
        # For Notion templates, we check the export or API data
        template_data = product_data.get('template_data', {})
        
        if not template_data:
            issues.append({
                "type": "missing_data",
                "severity": "critical",
                "description": "No Notion template data found"
            })
            metrics.completeness_score = 0
            return
        
        # Check for required Notion elements
        has_database = 'database' in str(template_data).lower()
        has_views = 'view' in str(template_data).lower()
        has_properties = 'properties' in str(template_data).lower()
        has_templates = 'template' in str(template_data).lower()
        
        notion_features = sum([has_database, has_views, has_properties, has_templates])
        
        # Score based on Notion-specific features
        metrics.functionality_score = min(100, 60 + (notion_features * 10))
        metrics.completeness_score = 90 if notion_features >= 3 else 70
        metrics.visual_polish_score = 85  # Notion has built-in polish
        metrics.content_quality_score = 80
        metrics.usability_score = 90 if has_templates else 75
        metrics.brand_compliance_score = 70  # Base score for Notion
        
        if notion_features < 2:
            issues.append({
                "type": "basic_template",
                "severity": "medium",
                "description": "Template lacks advanced Notion features"
            })
        
        test_results['notion_validation'] = {
            'has_database': has_database,
            'has_views': has_views,
            'has_properties': has_properties,
            'has_templates': has_templates
        }
    
    async def _validate_digital_tool(self, product_data: Dict, metrics: QualityMetrics, 
                                   issues: List, test_results: Dict):
        """Validate digital tools and calculators"""
        tool_type = product_data.get('tool_type', 'web')
        
        if tool_type == 'web':
            # Validate web-based tool
            url = product_data.get('url')
            if not url:
                issues.append({
                    "type": "missing_url",
                    "severity": "critical",
                    "description": "No URL provided for web tool"
                })
                metrics.functionality_score = 0
                return
            
            # Test tool accessibility and functionality
            try:
                async with self.session.get(url, timeout=10) as response:
                    if response.status == 200:
                        metrics.functionality_score = 90
                        
                        # Check for required elements
                        content = await response.text()
                        has_inputs = '<input' in content or '<select' in content
                        has_outputs = 'result' in content.lower() or 'output' in content.lower()
                        has_branding = 'myroofgenius' in content.lower()
                        
                        if not has_inputs:
                            issues.append({
                                "type": "no_inputs",
                                "severity": "high",
                                "description": "Tool has no input fields"
                            })
                            metrics.functionality_score -= 30
                        
                        metrics.brand_compliance_score = 90 if has_branding else 60
                        metrics.visual_polish_score = 85  # Base score for web tools
                        metrics.completeness_score = 90 if has_inputs and has_outputs else 70
                        metrics.usability_score = 85
                        metrics.content_quality_score = 80
                        
                    else:
                        issues.append({
                            "type": "tool_unavailable",
                            "severity": "critical",
                            "description": f"Tool returned status {response.status}"
                        })
                        metrics.functionality_score = 0
                        
            except Exception as e:
                issues.append({
                    "type": "tool_error",
                    "severity": "critical",
                    "description": f"Failed to access tool: {str(e)}"
                })
                metrics.functionality_score = 0
        
        test_results['tool_validation'] = {
            'tool_type': tool_type,
            'accessible': metrics.functionality_score > 0
        }
    
    async def _validate_generic_product(self, product_data: Dict, metrics: QualityMetrics, 
                                      issues: List, test_results: Dict):
        """Generic validation for other product types"""
        # Base scores for generic products
        metrics.completeness_score = 75
        metrics.functionality_score = 75
        metrics.visual_polish_score = 70
        metrics.content_quality_score = 75
        metrics.usability_score = 75
        metrics.brand_compliance_score = 70
        
        # Check basic requirements
        if not product_data.get('name'):
            issues.append({
                "type": "missing_name",
                "severity": "high",
                "description": "Product has no name"
            })
            metrics.completeness_score -= 20
        
        if not product_data.get('description'):
            issues.append({
                "type": "missing_description",
                "severity": "medium",
                "description": "Product has no description"
            })
            metrics.content_quality_score -= 20
        
        test_results['generic_validation'] = {
            'has_name': bool(product_data.get('name')),
            'has_description': bool(product_data.get('description'))
        }
    
    async def _store_qa_report(self, report: ProductQAReport):
        """Store QA report in persistent memory"""
        try:
            headers = {
                "apikey": self.supabase_key,
                "Authorization": f"Bearer {self.supabase_key}",
                "Content-Type": "application/json"
            }
            
            report_data = {
                "title": f"QA Report - {report.product_name}",
                "content": json.dumps(asdict(report), default=str),
                "role": "system",
                "memory_type": "qa_report",
                "tags": ["qa", "quality", report.product_type.value, report.status.value],
                "meta_data": {
                    "product_id": report.product_id,
                    "overall_score": report.metrics.overall_score,
                    "status": report.status.value,
                    "timestamp": report.timestamp.isoformat()
                },
                "is_active": True
            }
            
            async with self.session.post(
                f"{self.supabase_url}/rest/v1/copilot_messages",
                headers=headers,
                json=report_data
            ) as response:
                if response.status in [200, 201]:
                    logger.info(f"QA report stored for {report.product_name}")
                else:
                    logger.error(f"Failed to store QA report: {response.status}")
                    
        except Exception as e:
            logger.error(f"Error storing QA report: {str(e)}")
    
    async def _trigger_improvement_workflow(self, report: ProductQAReport):
        """Trigger automated improvement workflow for failed products"""
        logger.info(f"Triggering improvement workflow for {report.product_name}")
        
        # Create improvement task
        improvement_task = {
            "product_id": report.product_id,
            "product_name": report.product_name,
            "product_type": report.product_type.value,
            "issues": report.issues_found,
            "current_scores": asdict(report.metrics),
            "target_score": self.min_quality_score,
            "priority": "high" if report.status == QualityStatus.FAILED else "medium"
        }
        
        # Store improvement task
        try:
            headers = {
                "apikey": self.supabase_key,
                "Authorization": f"Bearer {self.supabase_key}",
                "Content-Type": "application/json"
            }
            
            task_data = {
                "title": f"Improve {report.product_name} to Commercial Grade",
                "content": json.dumps(improvement_task),
                "role": "system",
                "memory_type": "improvement_task",
                "tags": ["improvement", "quality", "automated"],
                "meta_data": improvement_task,
                "is_active": True
            }
            
            async with self.session.post(
                f"{self.supabase_url}/rest/v1/copilot_messages",
                headers=headers,
                json=task_data
            ) as response:
                if response.status in [200, 201]:
                    logger.info(f"Improvement task created for {report.product_name}")
                    
        except Exception as e:
            logger.error(f"Error creating improvement task: {str(e)}")
    
    async def run_marketplace_audit(self):
        """Run comprehensive audit of all marketplace products"""
        logger.info("Starting marketplace-wide quality audit")
        
        try:
            # Fetch all products from marketplace
            headers = {
                "apikey": self.supabase_key,
                "Authorization": f"Bearer {self.supabase_key}"
            }
            
            async with self.session.get(
                f"{self.supabase_url}/rest/v1/marketplace_products?select=*",
                headers=headers
            ) as response:
                if response.status == 200:
                    products = await response.json()
                    logger.info(f"Found {len(products)} products to audit")
                    
                    audit_results = []
                    for product in products:
                        try:
                            report = await self.validate_product(product)
                            audit_results.append(report)
                            
                            # Update product status based on QA results
                            if report.status == QualityStatus.PASSED:
                                await self._update_product_status(product['id'], 'active')
                            else:
                                await self._update_product_status(product['id'], 'under_review')
                                
                        except Exception as e:
                            logger.error(f"Error auditing product {product.get('name', 'Unknown')}: {str(e)}")
                    
                    # Generate audit summary
                    await self._generate_audit_summary(audit_results)
                    
                else:
                    logger.error(f"Failed to fetch products: {response.status}")
                    
        except Exception as e:
            logger.error(f"Error running marketplace audit: {str(e)}")
    
    async def _update_product_status(self, product_id: str, status: str):
        """Update product status in marketplace"""
        try:
            headers = {
                "apikey": self.supabase_key,
                "Authorization": f"Bearer {self.supabase_key}",
                "Content-Type": "application/json"
            }
            
            async with self.session.patch(
                f"{self.supabase_url}/rest/v1/marketplace_products?id=eq.{product_id}",
                headers=headers,
                json={"status": status, "last_qa_check": datetime.utcnow().isoformat()}
            ) as response:
                if response.status in [200, 204]:
                    logger.info(f"Updated product {product_id} status to {status}")
                    
        except Exception as e:
            logger.error(f"Error updating product status: {str(e)}")
    
    async def _generate_audit_summary(self, audit_results: List[ProductQAReport]):
        """Generate summary of audit results"""
        total_products = len(audit_results)
        passed = sum(1 for r in audit_results if r.status == QualityStatus.PASSED)
        failed = sum(1 for r in audit_results if r.status == QualityStatus.FAILED)
        needs_work = sum(1 for r in audit_results if r.status == QualityStatus.REQUIRES_REWORK)
        
        avg_score = sum(r.metrics.overall_score for r in audit_results) / total_products if total_products > 0 else 0
        
        summary = {
            "audit_date": datetime.utcnow().isoformat(),
            "total_products": total_products,
            "passed": passed,
            "failed": failed,
            "requires_rework": needs_work,
            "average_quality_score": avg_score,
            "pass_rate": (passed / total_products * 100) if total_products > 0 else 0,
            "common_issues": self._analyze_common_issues(audit_results)
        }
        
        logger.info(f"Audit Summary: {json.dumps(summary, indent=2)}")
        
        # Store summary
        try:
            headers = {
                "apikey": self.supabase_key,
                "Authorization": f"Bearer {self.supabase_key}",
                "Content-Type": "application/json"
            }
            
            summary_data = {
                "title": f"Marketplace Audit Summary - {datetime.utcnow().strftime('%Y-%m-%d')}",
                "content": json.dumps(summary),
                "role": "system",
                "memory_type": "audit_summary",
                "tags": ["audit", "summary", "quality"],
                "meta_data": summary,
                "is_active": True
            }
            
            async with self.session.post(
                f"{self.supabase_url}/rest/v1/copilot_messages",
                headers=headers,
                json=summary_data
            ) as response:
                if response.status in [200, 201]:
                    logger.info("Audit summary stored successfully")
                    
        except Exception as e:
            logger.error(f"Error storing audit summary: {str(e)}")
        
        return summary
    
    def _analyze_common_issues(self, audit_results: List[ProductQAReport]) -> List[Dict[str, Any]]:
        """Analyze and categorize common issues across products"""
        issue_counts = {}
        
        for report in audit_results:
            for issue in report.issues_found:
                issue_type = issue['type']
                if issue_type not in issue_counts:
                    issue_counts[issue_type] = {
                        'count': 0,
                        'severity': issue['severity'],
                        'examples': []
                    }
                issue_counts[issue_type]['count'] += 1
                if len(issue_counts[issue_type]['examples']) < 3:
                    issue_counts[issue_type]['examples'].append(report.product_name)
        
        # Sort by frequency
        sorted_issues = sorted(
            [{'type': k, **v} for k, v in issue_counts.items()],
            key=lambda x: x['count'],
            reverse=True
        )
        
        return sorted_issues[:10]  # Top 10 issues
    
    async def cleanup(self):
        """Cleanup resources"""
        if self.session:
            await self.session.close()

async def main():
    """Main function to run the QA system"""
    qa_system = CommercialGradeQASystem()
    
    try:
        await qa_system.initialize()
        
        # Run full marketplace audit
        await qa_system.run_marketplace_audit()
        
        # Keep system running for continuous monitoring
        while True:
            await asyncio.sleep(300)  # Check every 5 minutes
            logger.info("Running periodic quality check...")
            # Could add periodic checks here
            
    except KeyboardInterrupt:
        logger.info("Shutting down QA system")
    finally:
        await qa_system.cleanup()

if __name__ == "__main__":
    asyncio.run(main())
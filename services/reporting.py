# WeatherCraft ERP - Reporting Engine
# Comprehensive business intelligence and reporting

import logging
from datetime import datetime, timedelta, date
from typing import Dict, Any, List, Optional
import json
from decimal import Decimal

from sqlalchemy import text
from sqlalchemy.orm import Session
import pandas as pd

logger = logging.getLogger(__name__)

class ReportingEngine:
    """Comprehensive reporting and analytics engine"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ============================================================================
    # FINANCIAL REPORTS
    # ============================================================================
    
    def generate_profit_loss(self, start_date: date, end_date: date) -> Dict:
        """Generate profit and loss statement"""
        try:
            # Revenue
            revenue_query = """
            SELECT 
                SUM(total_amount) as total_revenue,
                COUNT(*) as invoice_count
            FROM invoices
            WHERE invoice_date BETWEEN :start_date AND :end_date
            AND status IN ('paid', 'partial')
            """
            
            revenue_result = self.db.execute(text(revenue_query), {
                'start_date': start_date,
                'end_date': end_date
            }).fetchone()
            
            # Costs
            cost_query = """
            SELECT 
                SUM(jc.total_cost) as total_costs,
                jc.cost_type,
                COUNT(DISTINCT jc.job_id) as job_count
            FROM job_costs jc
            JOIN jobs j ON jc.job_id = j.id
            WHERE j.completed_date BETWEEN :start_date AND :end_date
            GROUP BY jc.cost_type
            """
            
            cost_result = self.db.execute(text(cost_query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            costs_by_type = {}
            total_costs = 0
            
            for row in cost_result:
                costs_by_type[row['cost_type']] = float(row['total_costs'] or 0)
                total_costs += float(row['total_costs'] or 0)
            
            # Calculate metrics
            total_revenue = float(revenue_result['total_revenue'] or 0)
            gross_profit = total_revenue - total_costs
            gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
            
            return {
                'period': {
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'revenue': {
                    'total': total_revenue,
                    'invoice_count': revenue_result['invoice_count']
                },
                'costs': {
                    'total': total_costs,
                    'by_type': costs_by_type
                },
                'profit': {
                    'gross_profit': gross_profit,
                    'gross_margin': round(gross_margin, 2)
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating P&L: {e}")
            raise
    
    def generate_cash_flow(self, start_date: date, end_date: date) -> Dict:
        """Generate cash flow report"""
        try:
            # Cash inflows
            inflow_query = """
            SELECT 
                DATE_TRUNC('month', payment_date) as month,
                SUM(amount) as cash_in
            FROM payments
            WHERE payment_date BETWEEN :start_date AND :end_date
            AND status = 'completed'
            GROUP BY DATE_TRUNC('month', payment_date)
            ORDER BY month
            """
            
            inflow_result = self.db.execute(text(inflow_query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            # Cash outflows (vendor payments)
            outflow_query = """
            SELECT 
                DATE_TRUNC('month', payment_date) as month,
                SUM(amount_paid) as cash_out
            FROM vendor_invoices
            WHERE payment_date BETWEEN :start_date AND :end_date
            GROUP BY DATE_TRUNC('month', payment_date)
            ORDER BY month
            """
            
            outflow_result = self.db.execute(text(outflow_query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            # Combine results
            cash_flow = {}
            
            for row in inflow_result:
                month = str(row['month'].date())
                cash_flow[month] = {
                    'inflow': float(row['cash_in'] or 0),
                    'outflow': 0,
                    'net': float(row['cash_in'] or 0)
                }
            
            for row in outflow_result:
                month = str(row['month'].date())
                if month in cash_flow:
                    cash_flow[month]['outflow'] = float(row['cash_out'] or 0)
                    cash_flow[month]['net'] = cash_flow[month]['inflow'] - cash_flow[month]['outflow']
                else:
                    cash_flow[month] = {
                        'inflow': 0,
                        'outflow': float(row['cash_out'] or 0),
                        'net': -float(row['cash_out'] or 0)
                    }
            
            # Calculate totals
            total_inflow = sum(cf['inflow'] for cf in cash_flow.values())
            total_outflow = sum(cf['outflow'] for cf in cash_flow.values())
            net_cash_flow = total_inflow - total_outflow
            
            return {
                'period': {
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'monthly_cash_flow': cash_flow,
                'totals': {
                    'total_inflow': total_inflow,
                    'total_outflow': total_outflow,
                    'net_cash_flow': net_cash_flow
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating cash flow: {e}")
            raise
    
    def generate_ar_aging_report(self) -> Dict:
        """Generate accounts receivable aging report"""
        try:
            query = """
            SELECT 
                c.name as customer_name,
                c.id as customer_id,
                i.invoice_number,
                i.invoice_date,
                i.due_date,
                i.total_amount,
                i.balance_due,
                CURRENT_DATE - i.due_date as days_overdue,
                CASE 
                    WHEN i.due_date >= CURRENT_DATE THEN 'current'
                    WHEN CURRENT_DATE - i.due_date <= 30 THEN '1-30'
                    WHEN CURRENT_DATE - i.due_date <= 60 THEN '31-60'
                    WHEN CURRENT_DATE - i.due_date <= 90 THEN '61-90'
                    ELSE '90+'
                END as aging_bucket
            FROM invoices i
            JOIN customers c ON i.customer_id = c.id
            WHERE i.balance_due > 0
            AND i.status NOT IN ('paid', 'cancelled')
            ORDER BY days_overdue DESC
            """
            
            result = self.db.execute(text(query))
            invoices = [dict(row) for row in result]
            
            # Summarize by aging bucket
            aging_summary = {
                'current': {'count': 0, 'amount': 0},
                '1-30': {'count': 0, 'amount': 0},
                '31-60': {'count': 0, 'amount': 0},
                '61-90': {'count': 0, 'amount': 0},
                '90+': {'count': 0, 'amount': 0}
            }
            
            for invoice in invoices:
                bucket = invoice['aging_bucket']
                aging_summary[bucket]['count'] += 1
                aging_summary[bucket]['amount'] += float(invoice['balance_due'] or 0)
            
            total_ar = sum(bucket['amount'] for bucket in aging_summary.values())
            
            return {
                'generated_at': datetime.now().isoformat(),
                'summary': aging_summary,
                'total_ar': total_ar,
                'invoice_count': len(invoices),
                'details': invoices[:100]  # Limit to first 100 for performance
            }
            
        except Exception as e:
            logger.error(f"Error generating AR aging: {e}")
            raise
    
    # ============================================================================
    # OPERATIONAL REPORTS
    # ============================================================================
    
    def generate_job_profitability_report(self, start_date: date, end_date: date) -> Dict:
        """Generate job profitability analysis"""
        try:
            query = """
            SELECT 
                j.job_number,
                j.title,
                c.name as customer_name,
                j.total_amount as revenue,
                COALESCE(SUM(jc.total_cost), 0) as total_cost,
                j.total_amount - COALESCE(SUM(jc.total_cost), 0) as gross_profit,
                CASE 
                    WHEN j.total_amount > 0 THEN 
                        ((j.total_amount - COALESCE(SUM(jc.total_cost), 0)) / j.total_amount) * 100
                    ELSE 0 
                END as margin_percent,
                j.status,
                j.completed_date
            FROM jobs j
            LEFT JOIN customers c ON j.customer_id = c.id
            LEFT JOIN job_costs jc ON j.id = jc.job_id
            WHERE j.created_at BETWEEN :start_date AND :end_date
            GROUP BY j.id, j.job_number, j.title, c.name, j.total_amount, j.status, j.completed_date
            ORDER BY gross_profit DESC
            """
            
            result = self.db.execute(text(query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            jobs = []
            total_revenue = 0
            total_cost = 0
            
            for row in result:
                job = dict(row)
                job['revenue'] = float(job['revenue'] or 0)
                job['total_cost'] = float(job['total_cost'] or 0)
                job['gross_profit'] = float(job['gross_profit'] or 0)
                job['margin_percent'] = round(float(job['margin_percent'] or 0), 2)
                
                jobs.append(job)
                total_revenue += job['revenue']
                total_cost += job['total_cost']
            
            total_profit = total_revenue - total_cost
            avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
            
            return {
                'period': {
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'summary': {
                    'total_jobs': len(jobs),
                    'total_revenue': total_revenue,
                    'total_cost': total_cost,
                    'total_profit': total_profit,
                    'average_margin': round(avg_margin, 2)
                },
                'jobs': jobs[:50]  # Top 50 jobs
            }
            
        except Exception as e:
            logger.error(f"Error generating job profitability: {e}")
            raise
    
    def generate_crew_productivity_report(self, start_date: date, end_date: date) -> Dict:
        """Generate crew productivity analysis"""
        try:
            query = """
            SELECT 
                cr.crew_name,
                cr.id as crew_id,
                COUNT(DISTINCT s.id) as jobs_completed,
                SUM(te.total_hours) as total_hours,
                COUNT(DISTINCT te.user_id) as crew_size,
                AVG(qc.overall_score) as quality_score,
                SUM(j.total_amount) as revenue_generated
            FROM crews cr
            LEFT JOIN schedules s ON cr.id = s.crew_id
            LEFT JOIN time_entries te ON cr.id = te.crew_id
            LEFT JOIN jobs j ON s.schedulable_id = j.id AND s.schedulable_type = 'job'
            LEFT JOIN quality_checklists qc ON j.id = qc.job_id
            WHERE s.scheduled_date BETWEEN :start_date AND :end_date
            AND s.status = 'completed'
            GROUP BY cr.id, cr.crew_name
            ORDER BY revenue_generated DESC
            """
            
            result = self.db.execute(text(query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            crews = []
            for row in result:
                crew = dict(row)
                crew['total_hours'] = float(crew['total_hours'] or 0)
                crew['revenue_generated'] = float(crew['revenue_generated'] or 0)
                crew['quality_score'] = float(crew['quality_score'] or 0)
                
                # Calculate productivity metrics
                if crew['total_hours'] > 0:
                    crew['revenue_per_hour'] = crew['revenue_generated'] / crew['total_hours']
                else:
                    crew['revenue_per_hour'] = 0
                
                crews.append(crew)
            
            return {
                'period': {
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'crews': crews,
                'summary': {
                    'total_crews': len(crews),
                    'total_jobs': sum(c['jobs_completed'] for c in crews),
                    'total_hours': sum(c['total_hours'] for c in crews),
                    'total_revenue': sum(c['revenue_generated'] for c in crews)
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating crew productivity: {e}")
            raise
    
    def generate_material_usage_report(self, start_date: date, end_date: date) -> Dict:
        """Generate material usage and cost analysis"""
        try:
            query = """
            SELECT 
                ii.name as material_name,
                ii.item_code,
                SUM(it.quantity) as total_used,
                ii.unit_of_measure,
                AVG(it.unit_cost) as avg_cost,
                SUM(it.total_cost) as total_cost,
                COUNT(DISTINCT it.reference_id) as jobs_used
            FROM inventory_transactions it
            JOIN inventory_items ii ON it.item_id = ii.id
            WHERE it.transaction_type = 'issue'
            AND it.transaction_date BETWEEN :start_date AND :end_date
            GROUP BY ii.id, ii.name, ii.item_code, ii.unit_of_measure
            ORDER BY total_cost DESC
            """
            
            result = self.db.execute(text(query), {
                'start_date': start_date,
                'end_date': end_date
            })
            
            materials = []
            total_material_cost = 0
            
            for row in result:
                material = dict(row)
                material['total_used'] = float(material['total_used'] or 0)
                material['avg_cost'] = float(material['avg_cost'] or 0)
                material['total_cost'] = float(material['total_cost'] or 0)
                
                materials.append(material)
                total_material_cost += material['total_cost']
            
            return {
                'period': {
                    'start_date': str(start_date),
                    'end_date': str(end_date)
                },
                'materials': materials[:50],  # Top 50 materials
                'summary': {
                    'total_materials': len(materials),
                    'total_cost': total_material_cost,
                    'average_cost_per_material': total_material_cost / len(materials) if materials else 0
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating material usage: {e}")
            raise
    
    # ============================================================================
    # SALES REPORTS
    # ============================================================================
    
    def generate_sales_pipeline_report(self) -> Dict:
        """Generate sales pipeline analysis"""
        try:
            # Leads by stage
            leads_query = """
            SELECT 
                status,
                COUNT(*) as count,
                SUM(estimated_value) as total_value,
                AVG(score) as avg_score
            FROM leads
            WHERE status != 'closed'
            GROUP BY status
            """
            
            leads_result = self.db.execute(text(leads_query))
            
            # Estimates by status
            estimates_query = """
            SELECT 
                status,
                COUNT(*) as count,
                SUM(total_amount) as total_value
            FROM estimates
            WHERE created_at >= CURRENT_DATE - INTERVAL '90 days'
            GROUP BY status
            """
            
            estimates_result = self.db.execute(text(estimates_query))
            
            # Conversion metrics
            conversion_query = """
            SELECT 
                COUNT(DISTINCT l.id) as total_leads,
                COUNT(DISTINCT e.id) as total_estimates,
                COUNT(DISTINCT j.id) as total_jobs,
                COUNT(DISTINCT CASE WHEN j.status = 'completed' THEN j.id END) as completed_jobs
            FROM leads l
            LEFT JOIN estimates e ON l.id = e.lead_id
            LEFT JOIN jobs j ON e.id = j.estimate_id
            WHERE l.created_at >= CURRENT_DATE - INTERVAL '90 days'
            """
            
            conversion_result = self.db.execute(text(conversion_query)).fetchone()
            
            # Build pipeline summary
            pipeline = {
                'leads': {},
                'estimates': {},
                'conversion_rates': {}
            }
            
            for row in leads_result:
                pipeline['leads'][row['status']] = {
                    'count': row['count'],
                    'value': float(row['total_value'] or 0),
                    'avg_score': float(row['avg_score'] or 0)
                }
            
            for row in estimates_result:
                pipeline['estimates'][row['status']] = {
                    'count': row['count'],
                    'value': float(row['total_value'] or 0)
                }
            
            # Calculate conversion rates
            if conversion_result['total_leads'] > 0:
                pipeline['conversion_rates']['lead_to_estimate'] = (
                    conversion_result['total_estimates'] / conversion_result['total_leads'] * 100
                )
            else:
                pipeline['conversion_rates']['lead_to_estimate'] = 0
            
            if conversion_result['total_estimates'] > 0:
                pipeline['conversion_rates']['estimate_to_job'] = (
                    conversion_result['total_jobs'] / conversion_result['total_estimates'] * 100
                )
            else:
                pipeline['conversion_rates']['estimate_to_job'] = 0
            
            if conversion_result['total_jobs'] > 0:
                pipeline['conversion_rates']['job_completion'] = (
                    conversion_result['completed_jobs'] / conversion_result['total_jobs'] * 100
                )
            else:
                pipeline['conversion_rates']['job_completion'] = 0
            
            return {
                'generated_at': datetime.now().isoformat(),
                'pipeline': pipeline,
                'totals': {
                    'leads': conversion_result['total_leads'],
                    'estimates': conversion_result['total_estimates'],
                    'jobs': conversion_result['total_jobs'],
                    'completed': conversion_result['completed_jobs']
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating sales pipeline: {e}")
            raise
    
    def generate_customer_lifetime_value_report(self) -> Dict:
        """Generate customer lifetime value analysis"""
        try:
            query = """
            SELECT 
                c.id,
                c.name,
                c.customer_type,
                COUNT(DISTINCT j.id) as total_jobs,
                SUM(j.total_amount) as lifetime_revenue,
                AVG(j.total_amount) as avg_job_value,
                MIN(j.created_at) as first_job_date,
                MAX(j.created_at) as last_job_date,
                COUNT(DISTINCT st.id) as service_tickets,
                AVG(j.satisfaction_score) as avg_satisfaction
            FROM customers c
            LEFT JOIN jobs j ON c.id = j.customer_id
            LEFT JOIN service_tickets st ON c.id = st.customer_id
            WHERE j.status = 'completed'
            GROUP BY c.id, c.name, c.customer_type
            HAVING COUNT(DISTINCT j.id) > 0
            ORDER BY lifetime_revenue DESC
            LIMIT 100
            """
            
            result = self.db.execute(text(query))
            
            customers = []
            for row in result:
                customer = dict(row)
                customer['lifetime_revenue'] = float(customer['lifetime_revenue'] or 0)
                customer['avg_job_value'] = float(customer['avg_job_value'] or 0)
                customer['avg_satisfaction'] = float(customer['avg_satisfaction'] or 0)
                
                # Calculate customer tenure in days
                if customer['first_job_date'] and customer['last_job_date']:
                    tenure = (customer['last_job_date'] - customer['first_job_date']).days
                    customer['tenure_days'] = tenure
                else:
                    customer['tenure_days'] = 0
                
                customers.append(customer)
            
            # Calculate summary metrics
            total_customers = len(customers)
            total_revenue = sum(c['lifetime_revenue'] for c in customers)
            avg_ltv = total_revenue / total_customers if total_customers > 0 else 0
            
            return {
                'generated_at': datetime.now().isoformat(),
                'top_customers': customers,
                'summary': {
                    'total_customers': total_customers,
                    'total_lifetime_revenue': total_revenue,
                    'average_ltv': avg_ltv,
                    'average_jobs_per_customer': sum(c['total_jobs'] for c in customers) / total_customers if total_customers > 0 else 0
                }
            }
            
        except Exception as e:
            logger.error(f"Error generating CLV report: {e}")
            raise
    
    # ============================================================================
    # EXPORT FUNCTIONALITY
    # ============================================================================
    
    def export_to_excel(self, report_data: Dict, report_name: str) -> bytes:
        """Export report data to Excel format"""
        try:
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = report_name
            
            # Write report data
            row = 1
            for key, value in report_data.items():
                if isinstance(value, dict):
                    ws.cell(row=row, column=1, value=key)
                    row += 1
                    for sub_key, sub_value in value.items():
                        ws.cell(row=row, column=2, value=sub_key)
                        ws.cell(row=row, column=3, value=str(sub_value))
                        row += 1
                elif isinstance(value, list) and value:
                    # Write list as table
                    ws.cell(row=row, column=1, value=key)
                    row += 1
                    
                    # Write headers
                    if isinstance(value[0], dict):
                        headers = list(value[0].keys())
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=header)
                        row += 1
                        
                        # Write data
                        for item in value:
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=str(item.get(header, '')))
                            row += 1
                else:
                    ws.cell(row=row, column=1, value=key)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
                
                row += 1  # Add blank row between sections
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file.read()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise